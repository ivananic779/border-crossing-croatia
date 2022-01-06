import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import json, time, requests, re

month_dict = {
    'siječnja': '01.',
    'veljače': '02.',
    'ožujka': '03.',
    'travnja': '04.',
    'svibnja': '05.',
    'lipnja': '06.',
    'srpnja': '07.',
    'kolovoza': '08.',
    'rujna': '09.',
    'listopada': '10.',
    'studenog': '11.',
    'studenoga': '11.',
    'prosinca': '12.'
}

url = r'https://granica.mup.hr/default.inc.aspx?ajaxq=PrometPoDatumu&odDat='

base_date = datetime.today()
days_look_back = 20
date_list = [base_date - timedelta(days=x) for x in range(days_look_back)]

wb = load_workbook('template_ulazi_izlazi.xlsx')
sheet = wb.active

excel_row_number = 2
excel_column_number = 1

def get_html(_date):
    global url

    _url = url + _date.strftime('%d.%m.20%y.')

    try:
        r = requests.get(_url)
    except:
        # Try again, sometimes the first request fails
        try:            
            time.sleep(1)
            r = requests.get(_url)
        except Exception as e:
            print('Error get_html: ' + _url + ' ' + str(e))
            exit(1)

    return r.text

def get_reference_date(_html):
    """
        Find the date in the h2 tag, format it into a datetime object

        It sometimes appears as, where we want the second date:
        <h2 id='makniSe'>Objavljeno 17. studenog 2021. godine (podaci se odnose na prethodni dan, 16. studenog 2021.):</h2><div id='AjaxPoziv'>

        or as:
        <h2 id='makniSe'>Dohvaćeni su najkasnije dodani podaci od 5. siječnja 2022. godine:</h2><div id='AjaxPoziv'>

        Returns:
            datetime object
    """
    global month_dict

    # Read the h2 tag value, located in the first line
    h2_tag = _html.split('\n')[0]

    # Determine if the date is in the first or the second format
    if 'Objavljeno' in _html:
        # Use regex to read second date written as "5. siječnja 2022." "5. veljače 2021." etc.
        # from "<h2 id='makniSe'>Objavljeno 17. studenog 2021. godine (podaci se odnose na prethodni dan, 16. studenog 2021.):</h2><div id='AjaxPoziv'>"
        date_str = re.findall(r'\d+\.\s\w+\s\d+\.', h2_tag)

        # Check if date was found
        if len(date_str[1]) == 0:
            print('Error get_reference_date: date not found')
            exit(1)

        date_str = date_str[1]
    elif 'Dohvaćeni' in _html:
        # Use regex to read "5. siječnja 2022." "5. veljače 2021." etc.
        # from "<h2 id='makniSe'>Dohvaćeni su najkasnije dodani podaci od 5. siječnja 2022. godine:</h2><div id='AjaxPoziv'>"
        date_str = re.findall(r'\d+\.\s\w+\s\d+\.', h2_tag)

        # Check if date was found
        if len(date_str[0]) == 0:
            print('Error get_reference_date: date not found')
            exit(1)

        date_str = date_str[0]
    else:
        print('Error get_reference_date: h2 description not found')
        exit(1)

    try:
        # If the first number has 1 digit, add a 0 before it
        if len(date_str.split('.')[0]) == 1:
            date_str = '0' + date_str

        # Remove whitespace from date
        date_str = date_str.replace(' ', '')

        # Replace month written as words in date_str with string from month_dict
        # example date_str: "5.siječnja2022.", "5.veljače2021." etc.
        for key, value in month_dict.items():
            date_str = date_str.replace(key, value)

        # Convert date_str into datetime object
        date = datetime.strptime(date_str, '%d.%m.%Y.')
    except Exception as e:
        print('Error get_description_date: ' + str(e))
        exit(1)

    return date

def get_data(_html):
    """
        Returns list of dataFrames from the tables in html
    """
    try:
        data = pd.read_html(_html)
    except Exception as e:
        print('Error get_data: ' + str(e))
        exit(1)

    return data

def data_to_excel(_data, _date):
    global excel_column_number
    global excel_row_number

    for key, value in _data.items():
        try:
            if excel_column_number == 1:
                sheet.cell(row=excel_row_number,column=excel_column_number).value = _date
                excel_column_number += 1
            elif (key != '0'):
                sheet.cell(row=excel_row_number,column=excel_column_number).value = int(value)
                excel_column_number += 1
        except Exception as e:
            print('Error data_to_excel: ' + str(e))
            exit(1)


if __name__ == '__main__':

    i = 0

    for date in date_list:

        html = get_html(date)

        # Get the date written in the html
        reference_date = get_reference_date(html)

        # If reference date is the same as base_date - 1 day, then skip that date
        # Ignore time difference, just check if the date is the same
        if reference_date.date() == (base_date - timedelta(days=1)).date():
            continue

        data = get_data(html)

        json_putnici = json.loads(json.dumps(data[0].to_dict(orient='records'), indent=4))
        
        for row in json_putnici[3:9]:
            data_to_excel(row, (date - timedelta(days=1)).strftime('%d.%m.20%y.'))

        for row in json_putnici[12:]:
            data_to_excel(row, (date - timedelta(days=1)).strftime('%d.%m.20%y.'))

        # Increase row number for each date
        excel_row_number += 1
        # Reset column number for each date
        excel_column_number = 1

        i += 1

        print('Saved: ' + str(i))

    wb.save('ulazi_izlazi_01_01_2019.xlsx')
