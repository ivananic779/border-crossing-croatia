import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import json, time, requests, re
import pyodbc, os

conn = None

month_dict = {
    'siječnja': '01.',
    'veljače': '02.',
    'ožujka': '03.',
    'travnja': '04.',
    'svibnja': '05.',
    'lipnja': '06.',
    'srpnja': '07.',
    'kolovoza': '08.',
    'rujna': '09.',
    'listopada': '10.',
    'studenog': '11.',
    'studenoga': '11.',
    'prosinca': '12.'
}

table_dict = {
    0: 't_cestovni',
    1: 't_zeljeznicki',
    2: 't_rijecni',
    3: 't_pomorski',
    4: 't_zracni',
    5: 't_ukupno',
    6: 't_osobni_automobili',
    7: 't_vlakovi',
    8: 't_teretna_vozila',
    9: 't_autobusi',
    10: 't_plovila',
    11: 't_avioni',
    12: 't_ukupno_prijevozna_sredstva'
}

url = r'https://granica.mup.hr/default.inc.aspx?ajaxq=PrometPoDatumu&odDat='

base_date = datetime.today()
days_look_back = 4000
date_list = [base_date - timedelta(days=x) for x in range(days_look_back)]

wb = load_workbook('template_ulazi_izlazi.xlsx')
sheets = wb.sheetnames

sheet = wb[sheets[0]]

excel_row_number = 2
excel_column_number = 1

def connect_db():
    """
    Connect to the database

    Returns:
        connection: Cursor object
    """
    global conn

    # Get the database connection string from the environment variables
    server = 'DESKTOP-TBT2BKN\SQLEXPRESS'
    database = 'border_crossing'
    username = os.environ['DB_USERNAME']
    password = os.environ['DB_PASSWORD']

    # Connect to the database and return the connection
    try:
        conn = pyodbc.connect(
            'DRIVER={ODBC Driver 17 for SQL Server};'
            + 'SERVER=' + server
            + ';DATABASE=' + database
            + ';UID=' + username
            + ';PWD=' + password).cursor()

    except Exception as e:
        print(str(e))
        exit(1)

def get_html(_date):
    global url

    _url = url + _date.strftime('%d.%m.20%y.')

    try:
        r = requests.get(_url)
    except:
        # Try again, sometimes the first request fails
        try:            
            time.sleep(2)
            r = requests.get(_url)
        except Exception as e:
            print('Error get_html: ' + _url + ' ' + str(e))
            exit(1)

    return r.text

def get_reference_date(_html):
    """
        Find the date in the h2 tag, format it into a datetime object

        It sometimes appears as, where we want the second date:
        <h2 id='makniSe'>Objavljeno 17. studenog 2021. godine (podaci se odnose na prethodni dan, 16. studenog 2021.):</h2><div id='AjaxPoziv'>

        or as:
        <h2 id='makniSe'>Dohvaćeni su najkasnije dodani podaci od 5. siječnja 2022. godine:</h2><div id='AjaxPoziv'>

        Returns:
            datetime object
    """
    global month_dict

    # Read the h2 tag value, located in the first line
    h2_tag = _html.split('\n')[0]

    # Determine if the date is in the first or the second format
    if 'Objavljeno' in _html:
        # Use regex to read second date written as "5. siječnja 2022." "5. veljače 2021." etc.
        # from "<h2 id='makniSe'>Objavljeno 17. studenog 2021. godine (podaci se odnose na prethodni dan, 16. studenog 2021.):</h2><div id='AjaxPoziv'>"
        date_str = re.findall(r'\d+\.\s\w+\s\d+\.', h2_tag)

        # Check if date was found
        if len(date_str[1]) == 0:
            print('Error get_reference_date: date not found')
            exit(1)

        date_str = date_str[1]
    elif 'Dohvaćeni' in _html:
        # Use regex to read "5. siječnja 2022." "5. veljače 2021." etc.
        # from "<h2 id='makniSe'>Dohvaćeni su najkasnije dodani podaci od 5. siječnja 2022. godine:</h2><div id='AjaxPoziv'>"
        date_str = re.findall(r'\d+\.\s\w+\s\d+\.', h2_tag)

        # Check if date was found
        if len(date_str[0]) == 0:
            print('Error get_reference_date: date not found')
            exit(1)

        date_str = date_str[0]
    else:
        print('Error get_reference_date: h2 description not found')
        exit(1)

    try:
        # If the first number has 1 digit, add a 0 before it
        if len(date_str.split('.')[0]) == 1:
            date_str = '0' + date_str

        # Remove whitespace from date
        date_str = date_str.replace(' ', '')

        # Replace month written as words in date_str with string from month_dict
        # example date_str: "5.siječnja2022.", "5.veljače2021." etc.
        for key, value in month_dict.items():
            date_str = date_str.replace(key, value)

        # Convert date_str into datetime object
        date = datetime.strptime(date_str, '%d.%m.%Y.')
    except Exception as e:
        print('Error get_description_date: ' + str(e))
        exit(1)

    return date

def get_data(_html):
    """
        Returns list of dataFrames from the tables in html
    """
    try:
        data = pd.read_html(_html)
    except Exception as e:
        print('Error get_data: ' + str(e))
        exit(1)

    return data

def data_to_excel(_data, _date):
    """
        _data looks like this:
        {'0': 'Cestovni', '1': '46637', '2': '69364', '3': '116001', '4': '46979', '5': '57250', '6': '104229', '7': '220230'}
        from index 1 to index 7:
        domaci ulaz, strani ulaz, ukupno ulaz, domaci izlaz, strani izlaz, ukupno izlaz, ukupno

        shove the data into the excel file, each sheet is a different 0 index value(cestovni, zracni, automobili, avioni...)
        and they are always in the same order just like in the template
    """
    global excel_column_number
    global excel_row_number
    global sheet

    for key, value in _data.items():
        try:
            if excel_column_number == 1:
                sheet.cell(row=excel_row_number,column=excel_column_number).value = _date
                excel_column_number += 1
            elif (key != '0'):
                sheet.cell(row=excel_row_number,column=excel_column_number).value = int(value)
                excel_column_number += 1
        except Exception as e:
            print('Error data_to_excel: ' + str(e))
            exit(1)

def data_to_db(_data, _date, _table_index):
    insert_str = 'INSERT INTO ' + table_dict[_table_index] + ' (date, ulaz_domaci, ulaz_strani, ulaz_ukupno, izlaz_domaci, izlaz_strani, izlaz_ukupno, sveukupno) VALUES (?, ?, ?, ?, ?, ?, ?, ?)'
    
    try:
        conn.execute(insert_str, (_date, _data['1'], _data['2'], _data['3'], _data['4'], _data['5'], _data['6'], _data['7']))
    except Exception as e:
        conn.rollback()
        print('Error data_to_db: ' + str(e))
        exit(1)
    else:
        conn.commit()


if __name__ == '__main__':

    i = 0

    #connect_db()

    for date in date_list:

        html = get_html(date)

        # Get the date written in the html
        reference_date = get_reference_date(html)

        # If reference date is the same as base_date - 1 day, then skip that date
        # Ignore time difference, just check if the date is the same
        if reference_date.date() == base_date.date():
            with open('missing_dates.txt', 'a') as f:
                f.write(date.strftime('%d.%m.%Y.') + '\n')
            continue

        data = get_data(html)

        json_putnici = json.loads(json.dumps(data[0].to_dict(orient='records'), indent=4))

        # We will use this to change active sheets by referencing the index of the sheet in variable sheets
        # Each row in json_putnici is a new sheet
        j = 0
        
        for row in json_putnici[3:9]:
            sheet = wb[sheets[j]]
            data_to_excel(row, (date - timedelta(days=1)).strftime('%d.%m.20%y.'))
            #data_to_db(row, (date - timedelta(days=1)), j)
            j += 1
            # Reset column number for each sheet
            excel_column_number = 1

        for row in json_putnici[12:]:
            sheet = wb[sheets[j]]
            data_to_excel(row, (date - timedelta(days=1)).strftime('%d.%m.20%y.'))
            #data_to_db(row, (date - timedelta(days=1)), j)
            j += 1
            # Reset column number for each sheet
            excel_column_number = 1

        # Increase row number for each date
        excel_row_number += 1

        i += 1

        print('Saved: ' + str(i))

    wb.save(base_date.strftime('%d_%m_%Y_%H_%M') + '.xlsx')
